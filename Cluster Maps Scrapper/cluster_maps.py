import copy
import math
import random
import sys
import threading
import time

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from undetected_chromedriver import Chrome
from selenium.common.exceptions import TimeoutException

valid_mails = []
invalid_mails = []


def get_people():
    global valid_mails
    global invalid_mails
    chrome_options = Options()
    prefs = {"profile.default_content_setting_values.notifications": 2}
    chrome_options.add_experimental_option("prefs", prefs)
    # chrome_options.add_argument("--headless")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    if use_proxy:
        chrome_options.add_argument('--proxy-server=http://' + proxies[0])
    front_page = 'https://clustrmaps.com/d/'
    driver = None
    try:
        driver = webdriver.Chrome(service=Service('chromedriver.exe'), options=chrome_options)
    except (FileNotFoundError, Exception):
        print("Chromedriver not found")
        driver.close()
        exit(1)
    driver.get(front_page)
    state_ = BeautifulSoup(
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body'))).get_attribute(
            'innerHTML'), 'html.parser')
    states = state_.select('div.container.mt-5 > div.col-12 > div.row > div')
    error_str = 'State'
    for state_ in states:
        if state == state_.getText():
            error_str = 'City'
            link = 'https://clustrmaps.com' + state_.find('a').get('href')
            driver.get(link)
            city_table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR,
                                                                                         'body')))
            city_rows = BeautifulSoup(city_table.get_attribute('innerHTML'), 'html.parser')
            city_rows = city_rows.find(
                'div', {'class': 'table-responsive'})
            city_rows = city_rows.select('table > tbody > tr')
            for city_row in city_rows:
                city_tag = city_row.find_all('td')[1]
                city_name = city_tag.getText().strip().upper()
                if city == city_name:
                    driver.get('https://clustrmaps.com' + city_tag.find('a').get('href'))
                    fip_table = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, 'body')))
                    fip_html = BeautifulSoup(fip_table.get_attribute('innerHTML'), 'html.parser')
                    fip_rows = fip_html.select('div.col-12.mt-5')
                    for f in fip_rows:
                        f_ = f.find('h2')
                        if f_ is not None:
                            if 'FIPS' in f_.getText():
                                fip_rows = f.select('div')
                                break
                    error_str = 'FIP'
                    temp_people_links = []
                    for fip_row in fip_rows:
                        if fip_row.getText() in fip:
                            link = 'https://clustrmaps.com' + fip_row.find('a').get('href')
                            driver.get(link)
                            people_page = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
                            people_table = BeautifulSoup(people_page.get_attribute('innerHTML'), 'html.parser')
                            peoples_header = people_table.select_one('div.container.mt-4 > div > div:nth-child(7) > h2')
                            if peoples_header is None:
                                continue
                            peoples = people_table.select('div.container.mt-4 > div > div:nth-child(7) > div > div')
                            for people in peoples:
                                temp_people_links.append({people.find('a').getText(): 'https://clustrmaps.com' + people.find(
                                    'a').get('href')})
                    if len(temp_people_links) > 0:
                        driver.close()
                        return temp_people_links
                    break
            break
    print(error_str + " not found")
    driver.close()
    return []


def get_people_info(people: list, proxy_to_use: str):
    global index
    a = {}
    a.values()
    while True:
        try:
            chrome_options = Options()
            if use_proxy:
                chrome_options.add_argument('--proxy-server=http://' + proxy_to_use)
            driver = Chrome(options=chrome_options)
            driver.minimize_window()
            for links in people:
                index += 1
                j = index / len(people_links) * 100
                if j > 100:
                    j = 99
                    index -= 1
                progress(count=j, total=100)
                if list(links.keys())[0] in duplicates:
                    continue
                driver.get(list(links.values())[0])
                WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     'body > div:nth-child(2) > div > div:nth-child(1) > div > div:nth-child(1) > h1 > span')))
                html = BeautifulSoup(driver.find_element(By.CSS_SELECTOR, 'body').get_attribute('innerHTML'),
                                     'html.parser')
                name = html.find('span', {'itemprop': 'name'}).getText()
                age = html.find('small', {'class': 'text-muted'}).getText()
                int_age = ''
                if 'age' in age:
                    for char in age:
                        if char.isdigit():
                            int_age += char
                phone = html.find('h2', {'id': 'asPh'})
                p_str = []
                e_str = []
                if phone is not None:
                    phones = phone.parent.find_all('li')
                    for phone in phones:
                        if save_contacts_only:
                            if 'Landline' not in phone.getText():
                                p_str.append(phone.getText().split(',')[0])
                        else:
                            p_str.append(phone.getText().split(',')[0])
                            persons_associated = html.find('div', {'id': 'asPhPers'})
                            if persons_associated is not None:
                                card = persons_associated.find_all('div', {'class', 'flex-grow-1'})
                                for person in card:
                                    name = person.find('a').getText()
                                    if name.strip() not in duplicates:
                                        duplicates.append(name.strip())
                else:
                    continue
                if not save_contacts_only:
                    email = html.find('h2', {'id': 'asMail'})
                    if email is not None:
                        emails = email.parent.find_all('li')
                        for email in emails:
                            e_str.append(email.find('span').getText() + '\n')
                if len(p_str) > 1:
                    details.append([name, int_age, p_str, e_str])
                time.sleep(1)
            threads[int(threading.currentThread().name)] = False
            driver.close()
            return
        except (TimeoutException, Exception):
            while True and use_proxy:
                proxy_index = random.randint(0, len(proxies) - 1)
                if proxy_index not in proxy_indexes:
                    proxy_indexes.append(proxy_index)
                    proxy_to_use = proxies[proxy_index]
                    break


def write_to_file():
    progress(0, 100)
    global details
    details_index = 0
    row_index = 2
    global break_off
    while True:
        for detail in range(details_index, len(details)):
            sheet['A' + str(row_index)].value = details[detail][0]
            sheet['B' + str(row_index)].value = details[detail][1]
            col = 3
            for p in details[detail][2]:
                sheet[get_column_letter(col) + str(row_index)].value = p[0:14]
                col += 1
                sheet[get_column_letter(col) + str(row_index)].value = 'Landline' if 'Landline' in p else 'Cellphone'
                col += 1
                if col == 13:
                    break
            col = 13
            for e in details[detail][3]:
                sheet[get_column_letter(col) + str(row_index)].value = e
                col += 1
                if col == 18:
                    break
            row_index += 1
            book.save('people_info.xlsx')
            row_index += 1
            details_index += 1
        if break_off:
            break


def progress(count, total, suffix=''):
    sys.stdout.flush()  # As suggested by Rom Ruben
    bar_len = 100
    filled_len = int(round(bar_len * count / float(total)))

    percents = round(100.0 * count / float(total), 1)
    bar = '█' * filled_len + '-' * (bar_len - filled_len)
    sys.stdout.write('BOT PROGRESS : [%s] %s%s ...%s\r' % (bar, percents, '%', suffix))
    if percents == total:
        print()


def threaded_search(links_: list):
    thread_count = ((20 if len(links_) >= 20 else len(links_)) if use_threads else 1)
    no_of_inputs = math.floor(len(links_) / thread_count)
    for thread in range(thread_count):
        start = thread * no_of_inputs
        end = no_of_inputs * (thread + 1)
        if use_proxy:
            while True:
                proxy_index = random.randint(0, len(proxies) - 1)
                if proxy_index not in proxy_indexes:
                    proxy_indexes.append(proxy_index)
                    break
        if thread == thread_count - 1:
            end = len(links_)
        threading.Thread(target=get_people_info, name=str(thread), args=(
            copy.copy(links_[start:end]),
            (proxies[proxy_index] if use_proxy else None)
        )).start()
        threads.append(True)
    threading.Thread(target=write_to_file).start()
    global break_off
    while True:
        all_done = True
        for t in threads:
            if t:
                all_done = False
                break
        if all_done:
            break_off = True
            break


if __name__ == '__main__':
    break_off = False
    duplicates = []
    details = []
    index = 0
    book = openpyxl.load_workbook('people_info.xlsx')
    sheet = book.active
    while sheet.max_row > 1:
        sheet.delete_rows(2)
    try:
        book.save('people_info.xlsx')
    except PermissionError:
        print('Error : Excel file open in background!')
        exit(1)
    state = input('Please enter the state in US : ').strip().upper()
    city = input('Please enter city in ' + state + ' : ').strip().upper()
    fip = []
    with open('fip.txt', 'r') as fip_file:
        for fip_value in fip_file:
            fip.append(fip_value.strip())
    save_contacts_only = 'y' in input("Do you want to save cellphone leads only(y/n)?").lower()
    use_proxy = True
    use_threads = 'y' in input("Do you want to use threads(y/n)?").lower()
    if not use_threads:
        use_proxy = 'y' in input("Do you want to use proxy(y/n)?").lower()
    print('***********************************************************')
    print('Searching....')
    proxies = []
    proxy_indexes = []
    threads = []
    if use_proxy:
        proxy_file = open('proxies.txt', 'r')
        for proxy in proxy_file:
            proxies.append(proxy.strip())
        proxy_file.close()
    people_links = get_people()
    if len(people_links) > 0:
        print('***********************************************************')
        print('People gathered at FIP...')
        print('Getting info of ' + str(len(people_links)) + ' people......')
        print('This may take a while..')
        print("WARNING! Don't open the excel file.")
        threaded_search(people_links)
        empty_rows = []
        for i in range(2, sheet.max_row + 1):
            if sheet['A' + str(i)].value is None:
                empty_rows.append(sheet['A' + str(i)])
            else:
                phone = list(sheet['C' + str(i) + ':L' + str(i)][0])
                client_phones = []
                for j in range(len(phone)):
                    if j % 2 == 0 and phone[j].value is not None:
                        client_phones.append(phone[j].value)
                for j in range(i + 1, sheet.max_row + 1):
                    if sheet['A' + str(j)].value is not None:
                        phone_temp = list(sheet['C' + str(j) + ':L' + str(j)][0])
                        for k in range(len(phone_temp)):
                            if k % 2 == 0 and phone[k].value is not None:
                                if phone_temp[k].value in client_phones:
                                    empty_rows.append(sheet['A' + str(j)])
                                    break
        for empty_row in empty_rows:
            sheet.delete_rows(empty_row.row, 1)
        book.save('people_info.xlsx')
        print()
        print('***********************************************************')
        print('All Data Gathered!')
        print("You can now open the excel file.")
        print('***********************************************************')
    else:
        print('***********************************************************')
        print('No people found!')
        print('***********************************************************')
    input('Press any key to continue!')
