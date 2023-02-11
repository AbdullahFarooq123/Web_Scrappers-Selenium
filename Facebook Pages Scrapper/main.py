import datetime
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

from login import Login
from pages import GetPages
import json

credentials = json.load(open('credentials.json', 'r'))
if 'username' not in credentials:
    user_name = input("Please enter username of facebook : ")
    password = input("Please enter password of facebook : ")
else:
    user_name = credentials['username']
    password = credentials['password']
give_followers = 'y' in input('Do you want to give followers in this run(y/n)? : ')
give_likes_and_views = 'y' in input('Do you want to give likes in this run(y/n)? : ')
chrome_options = Options()
prefs = {"profile.default_content_setting_values.notifications": 2}
chrome_options.add_argument('--disable-notifications')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_experimental_option("prefs", prefs)
pages = []
book = load_workbook('fb_list.xlsx')
sheet = book['fb_pages']
driver = None
row = 2
col = 1
while True:
    link = sheet['A' + str(row)].value
    if link is None:
        break
    else:
        pages.append(link)
    row += 1

try:
    driver = webdriver.Chrome(service=Service('chromedriver.exe'), options=chrome_options)
    login_error = 0
    network_error = 0
    login = Login(driver=driver, username=user_name, password=password)
    wait = 10
    while True:
        try:
            login.signin_fb(wait)
            print("Login successful")
            if 'username' not in credentials:
                credentials['username'] = user_name
                credentials['password'] = password
                json.dump(credentials, open('credentials.json', 'w'), indent=4)
            break
        except TimeoutException:
            wait += 5
            network_error += 1
            if network_error > 5:
                print('Network Error')
                driver.close()
                exit(1)
        except Exception:
            login_error += 1
            if login_error > 2:
                print('Login error user-name or password-incorrect')
                driver.close()
                exit(1)
except (FileNotFoundError, Exception) as E:
    print("Chromedriver not found")
    driver.close()
    exit(1)
wait = 2
index = 0
error = 0
g = {index, error}
while True:
    pages = []
    book = load_workbook('fb_list.xlsx')
    sheet = book['fb_pages']
    row = 2
    col = 1
    while True:
        link = sheet['A' + str(row)].value
        if link is None:
            break
        else:
            pages.append(link)
        row += 1
    print("Processing bot!")
    new_day = False
    date = datetime.date.today()
    post_time_today = sheet['C1'].value
    today_col = 3
    while post_time_today is not None:
        if post_time_today == date.strftime('%x'):
            break
        else:
            today_col += 1
        post_time_today = sheet[get_column_letter(today_col) + '1'].value
    if post_time_today is None:
        sheet[get_column_letter(today_col) + '1'].value = date.strftime('%x')
        new_day = True
    while index < len(pages):
        get_pages = GetPages(driver, pages[index])
        time.sleep(wait)
        flag = get_pages.get_num_followers(wait)
        index += 1
        wait = 2
        print(
            str(index) + " : FOLLOWERS AT " + get_pages.url + " : " + str(get_pages.no_of_followers))
        if get_pages.get_stories(new_day=new_day):
            print(
                str(index) + " : NEW STORY AT " + get_pages.url)
        get_pages.get_post_link(get_pages.url, new_day, give_followers, give_likes_and_views)
        print(
            str(index) + " : NEW POSTS AT " + get_pages.url + " : " + str(get_pages.new_posts))
        if sheet['B' + str(index + 1)].value is None:
            sheet['B' + str(index + 1)].value = str(get_pages.no_of_followers) + '(' + str(
                get_pages.no_of_followers) + ')'
        else:
            values = sheet['B' + str(index + 1)].value
            if '(' in values:
                prev_followers = ''
                for i in range(values.find('(') + 1, values.find(')')):
                    prev_followers += values[i]
                values = values.replace(prev_followers, str(get_pages.no_of_followers))
                sheet['B' + str(index + 1)].value = values
            else:
                sheet['B' + str(index + 1)].value = str(get_pages.no_of_followers) + '(' + str(
                    get_pages.no_of_followers) + ')'
        previous_posts = sheet[get_column_letter(today_col) + str(index + 1)].value
        if previous_posts is None:
            previous_posts = 0
        else:
            previous_posts = int(previous_posts)
        sheet[get_column_letter(today_col) + str(index + 1)].value = previous_posts + get_pages.new_posts
    book.save('fb_list.xlsx')
    print("Bot slept for 5 minutes")
    wait = 2
    index = 0
    error = 0
    time.sleep(300)
