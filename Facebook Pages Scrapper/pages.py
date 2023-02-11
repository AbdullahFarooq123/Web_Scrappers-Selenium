import json
import math
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException

from Email import emailAlert


class GetPages:
    def __init__(self, driver: WebDriver, url: str):
        self.driver = driver
        self.url = url
        self.driver.get(url)
        self.no_of_followers = 0
        self.new_posts = 0

    def get_num_followers(self, wait):
        time.sleep(5)
        follower_btn = WebDriverWait(self.driver, wait).until(EC.presence_of_element_located((By.CSS_SELECTOR,
                                                                                              'body')))
        html = BeautifulSoup(follower_btn.get_attribute('innerHTML'), 'html.parser')
        containers = html.find_all('div', {'class': 'sjgh65i0'})
        followers = ''
        likes = ''
        for container in containers:
            container_data = container.getText()
            if 'Intro' in container_data or 'About' in container_data:
                rows = container.find_all('span', {
                    'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh a8c37x1j fe6kdd0r mau55g9w c8b282yb keod5gw0 nxhoafnm aigsh9s9 d3f4x2em iv3no6db jq4qci2q a3bd9o3v b1v8xokw oo9gr5id hzawbc8m'})
                for row in rows:
                    data = row.getText()
                    if 'people like this' in data:
                        likes = data
                    elif 'people follow this' in data:
                        followers = data
                    elif 'Followed by' in data:
                        followers = data
        if not len(followers) == 0:
            if ' people follow this' in followers:
                self.no_of_followers = int(followers.replace(' people follow this', '').replace(',', '').strip())
            else:
                self.no_of_followers = int(followers.replace('Followed by ', '').replace(',', '').replace(' people',
                                                                                                          '').strip())
        elif not len(likes) == 0:
            self.no_of_followers = int(followers.replace(' people like this', '').replace(',', '').strip())
        # else:
        #     top_likes = html.find('div', {
        #         'class': 'j83agx80 mpmpiqla ahl66waf tmq14sqq rux31ns4 sjcfkmk3 dti9y0u4 nyziof1z'}).find('span', {
        #         'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh a8c37x1j fe6kdd0r mau55g9w c8b282yb keod5gw0 nxhoafnm aigsh9s9 d3f4x2em mdeji52x a5q79mjw g1cxx5fr b1v8xokw m9osqain'})
        #     columns = top_likes.find_all('a', {
        #         'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gpro0wi8 m9osqain lrazzd5p'})
        #     for column in columns:
        #         data = column.getText()
        #         if 'followers' in data:
        #             data = data.replace(' followers', '').replace(',', '').strip()
        #             if 'K' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 3)
        #             elif 'M' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 6)
        #             else:
        #                 self.no_of_followers = int(data)
        #         elif 'likes' in data:
        #             data = data.replace('likes', '').replace(',', '').strip()
        #             if 'K' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 3)
        #             elif 'M' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 6)
        #             else:
        #                 self.no_of_followers = int(data)
        #         elif 'friends' in data:
        #             data = data.replace('friends', '').replace(',', '').strip()
        #             if 'K' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 3)
        #             elif 'M' in data.upper():
        #                 self.no_of_followers = int(float(data[:-1]) * 10 ** 6)
        #             else:
        #                 self.no_of_followers = int(data)

    def get_stories(self, new_day):
        self.driver.get(self.url)
        data_file = open('story.json', 'r')
        data = json.load(data_file)
        data_file.close()
        if self.url not in data or new_day:
            data[self.url] = {'time': 9999, 'no': 0}
        a_file = open("story.json", "w")
        json.dump(data, a_file, indent=4)
        a_file.close()
        time.sleep(2)
        body = BeautifulSoup(self.driver.find_element(By.CSS_SELECTOR, 'body').get_attribute('innerHTML'),
                             'html.parser')
        name = body.find('div', {'class': 'j83agx80 mpmpiqla ahl66waf tmq14sqq rux31ns4 sjcfkmk3 dti9y0u4 nyziof1z'})
        if name is not None:
            name = name.find('h1', {'class': 'gmql0nx0 l94mrbxd p1ri9a11 lzcic4wl'}).getText()
        time.sleep(2)
        profile = self.driver.find_elements(By.TAG_NAME, 'div')
        for p in profile:
            label = p.get_attribute('aria-label')
            if label is not None and ('Page profile photo' in label or (name in label) if name is not None else True):
                profile = p
                break
        if type(profile) is list:
            return False
        try:
            profile.click()
        except ElementClickInterceptedException:
            return False
        time.sleep(2)
        body = BeautifulSoup(self.driver.find_element(By.CSS_SELECTOR, 'body').get_attribute('innerHTML'),
                             'html.parser')
        time.sleep(1)
        body_text = body.getText()
        if 'View story' in body_text:
            link_to_stories = body.find('a', {
                'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz p7hjln8o esuyzwwr f1sip0of n00je7tq arfg74bv qs9ysxi8 k77z8yql abiwlrkh p8dawk7l lzcic4wl dwo3fsh8 rq0escxv nhd2j8a9 j83agx80 btwxx1t3 pfnyh3mw opuu4ng7 kj2yoqh6 kvgmc6g5 oygrvhab l9j0dhe7 i1ao9s8h du4w35lb bp9cbjyn cxgpxx05 dflh9lhu sj5x9vvc scb9dxdr'}).get(
                'href')
            self.driver.get(f'https://www.facebook.com{link_to_stories}')
            body = BeautifulSoup(WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'body'))).get_attribute('innerHTML'),
                                 'html.parser')
            time.sleep(5)
            post_time = body.find_all('div', {
                'class': 'oajrlxb2 rq0escxv p7hjln8o f1sip0of n00je7tq arfg74bv qs9ysxi8 k77z8yql l9j0dhe7 abiwlrkh lzcic4wl g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz gmql0nx0 nhd2j8a9 ihxqhq3m l94mrbxd aenfhxwr kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h myohyog2 bj9fd4vl ksdfmwjs gofk2cf1 a8c37x1j k4urcfbm tm8avpzi'})
            for p in post_time:
                if name in p.getText():
                    post_time = p.find('span', {'class',
                                                'd2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh a8c37x1j fe6kdd0r mau55g9w c8b282yb keod5gw0 nxhoafnm aigsh9s9 d3f4x2em iv3no6db jq4qci2q a3bd9o3v b1v8xokw m9osqain'}).getText()
                    break
            reversed_time = (post_time[::-1])
            time_only = ''
            one_digit = False
            for char in reversed_time:
                if char == 'm' or char == 'h' or (not one_digit and char == ' ') or char.isdigit():
                    if char.isdigit():
                        one_digit = True
                    time_only += char
                else:
                    break
            time_only = time_only[::-1]
            no_of_new = 0
            if 'new' in post_time:
                no_of_new = int(post_time.replace(f'new  · {time_only}', ''))
            if 'h' in time_only:
                digits = ''
                for digit in time_only:
                    if digit.isdigit():
                        digits += digit
                    else:
                        break
                mins = int(digits) * 60
                time_only = time_only.replace(digits, str(mins)).replace('h', 'm')
            time_only = int(time_only.replace(' m', ''))
            prev_time = int(data[self.url]['time'])
            prev_new = int(data[self.url]['no'])
            if prev_time > time_only or (no_of_new - prev_new) > 0:
                data[self.url] = {'time': time_only, 'no': no_of_new}
                emailAlert('Page Update',
                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that following link has a new story\nLink : {self.url}')
                a_file = open("story.json", "w")
                json.dump(data, a_file, indent=4)
                a_file.close()
                return True
            else:
                data[self.url] = {'time': time_only, 'no': no_of_new}
                a_file = open("story.json", "w")
                json.dump(data, a_file, indent=4)
                a_file.close()
                return False
        return False

    def get_post_link(self, page: str, new_day: bool, give_followers: bool, give_likes: bool):
        self.driver.get(self.url)
        data_file = open('data.json', 'r')
        data = json.load(data_file)
        new_page = page not in data
        if page not in data:
            data[page] = []
        data_file.close()
        current_page_link = data[page]
        new_links = []
        for i in range(0, 10000 if new_page else 3000, 1000):
            self.driver.execute_script('window.scrollBy(' + str(i) + ',' + str(i + 1000) + ')')
            time.sleep(2)
        time.sleep(5)
        posts = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
        html = BeautifulSoup(posts.get_attribute('innerHTML'), 'html.parser')
        posts = html.find_all('div', {'class': 'du4w35lb k4urcfbm l9j0dhe7 sjgh65i0'})
        link_count = 0
        for post in posts:
            image_post = post.find('a', {
                'class': 'oajrlxb2 gs1a9yip g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 mg4g778l pfnyh3mw p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb n00je7tq arfg74bv qs9ysxi8 k77z8yql btwxx1t3 abiwlrkh p8dawk7l lzcic4wl a8c37x1j tm8avpzi'
            })
            # link_post = post.find('a', {
            #     'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8 datstx6m k4urcfbm'
            # })
            video_post = post.find('a', {
                'class': 'oajrlxb2 rq0escxv p7hjln8o i1ao9s8h esuyzwwr f1sip0of n00je7tq arfg74bv qs9ysxi8 k77z8yql l9j0dhe7 abiwlrkh p8dawk7l g5ia77u1 gcieejh5 bn081pho humdl8nn izx4hr6d nhd2j8a9 q9uorilb jnigpg78 qjjbsfad fv0vnmcu w0hvl6rk ggphbty4 byekypgc jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i2p6rm4e lzcic4wl awjy3rxs'
            })
            multi_image_post = post.find('a', {
                'class': 'oajrlxb2 gs1a9yip g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 mg4g778l pfnyh3mw p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of du4w35lb n00je7tq arfg74bv qs9ysxi8 k77z8yql btwxx1t3 abiwlrkh p8dawk7l q9uorilb lzcic4wl i09qtzwb n7fi1qx3 pmk7jnqg j9ispegn kr520xx4 tm8avpzi'
            })
            if image_post is not None:
                link = image_post.get('href').split('__cft__')[0]
                if link in current_page_link:
                    link_count += 1
                    if link_count > 2:
                        break
                else:
                    self.new_posts += (0 if new_page else 1)
                    new_links.append(link)
            # elif link_post is not None:
            #     link = link_post.get('href')
            #
            #     if link in current_page_link:
            #         link_count += 1
            #         if link_count > 2:
            #             break
            #     else:
            #         self.new_posts += (0 if new_page else 1)
            #         new_links.append(link)
            elif video_post is not None:
                link = video_post.get('href').split('__cft__')[0]
                if link in current_page_link:
                    link_count += 1
                    if link_count > 2:
                        break
                else:
                    self.new_posts += (0 if new_page else 1)
                    new_links.append(link)
            elif multi_image_post is not None:
                link = multi_image_post.get('href')
                self.driver.get(link)
                post_link = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'body')))
                time.sleep(5)
                post_html = BeautifulSoup(post_link.get_attribute('innerHTML'), 'html.parser')
                link_div = post_html.find_all('div', {
                    'class': 'rq0escxv l9j0dhe7 du4w35lb j83agx80 cbu4d94t pfnyh3mw d2edcug0 hpfvmrgz ph5uu5jm b3onmgus e5nlhep0 ecm0bbzt'})[
                    1]
                for l_ in link_div:
                    link_a = l_.find('a', {
                        'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gpro0wi8 oo9gr5id lrazzd5p'
                    })
                    if link_a is not None:
                        link = link_a.get('href').split('__cft__')[0]
                        if link in current_page_link:
                            link_count += 1
                            if link_count > 2:
                                break
                        else:
                            self.new_posts += (0 if new_page else 1)
                            new_links.append(link)

        data[page] = data[page] + new_links
        a_file = open("data.json", "w")
        json.dump(data, a_file, indent=4)
        a_file.close()
        if not self.new_posts == 0:
            emailAlert('Page Update',
                       f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that following link has {self.new_posts} new posts! \nLink : {self.url}')
            self.order(new_links, new_day, give_followers, give_likes)

    def multiple_of_100(self, value: int):
        return math.floor(value / 100) * 100

    def get_max_row(self, sheet: Workbook, col_char: str):
        index = 2
        while True:
            if sheet[col_char + str(index)].value is not None:
                index += 1
            else:
                index -= 1
                break
        return str(index)

    def order(self, links: list, new_day: bool, give_followers: bool, give_likes: bool):
        if self.no_of_followers >= 15000:
            emailAlert('Page Update',
                       f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that following link has followers greater than or equal to 15k\nLink : {self.url}\nFollowers : {self.no_of_followers}')
        rules = self.get_rules()
        if len(rules) == 0:
            return
        likes = self.multiple_of_100(int(self.no_of_followers * float(rules['likes'].replace('%', '')) / 100))
        view_videos = self.multiple_of_100(
            int(self.no_of_followers * float(rules['view videos'].replace('%', '')) / 100))
        video_post_like = self.multiple_of_100(
            int(self.no_of_followers * float(rules['video posts like'].replace('%', '')) / 100))
        followers = int(open('followers.txt', 'r').read().strip())
        sheet = load_workbook('fb_list.xlsx')['smm_follows_ids']
        followers_id = sheet['A2'].value
        likes_id = list(_id[0].value for _id in sheet['B2':'B' + self.get_max_row(sheet, 'B')])
        views_id = list(_id[0].value for _id in sheet['C2':'C' + self.get_max_row(sheet, 'C')])
        while True:
            if new_day and give_followers:
                order = (requests.get(
                    f'https://smmfollows.com/api/v2?key=83a67d5b6906fd48f8ee2a7343630835&action=add&service={str(followers_id.split(":")[0])}&link={str(self.url)}&quantity={str(followers)}') if 'smm' in str(
                    followers_id.split(":")[1]) else requests.get(
                    f'https://paytosmm.com/api/v2?key=20069f39dcdd7aa14cecdb66afd03389&action=add&service={str(followers_id.split(":")[0])}&link={str(self.url)}&quantity={str(followers)}'))
                response = order.json()
                if 'error' in response:
                    emailAlert('Smm Error',
                               f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order followers for pages due to error : \n{response["error"]}')
                    if 'Not enough funds on balance' in response['error']:
                        input('Error Followers : Please recharge your account and press any key to continue')
                    else:
                        print('Error Followers : ' + response['error'])
                        input(f'Error Followers : At id {followers_id}Correct Error and press any key to continue')
                        sheet = load_workbook('fb_list.xlsx')['smm_follows_ids']
                        followers_id = sheet['A2'].value
                else:
                    break
            else:
                break
        likes_index = 0
        views_index = 0
        if give_likes:
            for link in links:
                if '/videos/' in link:
                    while True:
                        order = (requests.get(
                            f'https://smmfollows.com/api/v2?key=83a67d5b6906fd48f8ee2a7343630835&action=add&service={str(likes_id[likes_index].split(":")[0])}&link={str(link)}&quantity={str(int(video_post_like))}') if 'smm' in str(
                            likes_id[likes_index].split(":")[1]) else requests.get(
                            f'https://paytosmm.com/api/v2?key=20069f39dcdd7aa14cecdb66afd03389&action=add&service={str(likes_id[likes_index].split(":")[0])}&link={str(self.url)}&quantity={str(int(video_post_like))}'))
                        response = order.json()
                        if 'error' in response:
                            webpage = str(likes_id[likes_index].split(":")[1])
                            if 'Not enough funds on balance' in response['error']:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order likes for videos due to error from {webpage}: \n{response["error"]}')
                                input(
                                    f'Error Likes {webpage} : Please recharge your account and press any key to continue')
                            elif 'less than minimal' in response['error']:
                                get_minimal = int(response['error'].replace('Quantity less than minimal', '').strip())
                                video_post_like = get_minimal

                            else:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order likes for videos due to error : \n{response["error"]}')
                                print(f'Error Likes {webpage}: ' + response['error'])
                                input(
                                    f'Error Likes {webpage} : At id {likes_id[likes_index]}Correct Error and press any key to continue')
                                sheet = load_workbook('fb_list.xlsx')['smm_follows_ids']
                                likes_id = list(_id[0].value for _id in sheet['B2':'B' + self.get_max_row(sheet, 'B')])
                        else:
                            break

                    while True:
                        order = (requests.get(
                            f'https://paytosmm.com/api/v2?key=20069f39dcdd7aa14cecdb66afd03389&action=add&service={str(views_id[views_index].split(":")[0])}&link={str(link)}&quantity={str(int(view_videos))}') if 'smm' in str(
                            views_id[views_index].split(":")[1]) else requests.get(
                            f'https://paytosmm.com/api/v2?key=20069f39dcdd7aa14cecdb66afd03389&action=add&service={str(views_id[views_index].split(":")[0])}&link={str(link)}&quantity={str(int(view_videos))}'))
                        response = order.json()
                        if 'error' in response:
                            webpage = str(views_id[views_index].split(":")[1])
                            if 'Not enough funds on balance' in response['error']:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order views for videos due to error : \n{response["error"]}')
                                input(
                                    f'Error Views {webpage} : Please recharge your account and press any key to continue')
                            elif 'less than minimal' in response['error']:
                                get_minimal = int(response['error'].replace('Quantity less than minimal', '').strip())
                                view_videos = get_minimal
                            else:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order views for videos due to error : \n{response["error"]}')
                                print(f'Error Views {webpage} : ' + response['error'])
                                input(
                                    f'Error Likes {webpage} : At id {views_id[views_index]}Correct Error and press any key to continue')
                                sheet = load_workbook('fb_list.xlsx')['smm_follows_ids']
                                views_id = list(_id[0].value for _id in sheet['C2':'C' + self.get_max_row(sheet, 'C')])
                        else:
                            break
                    views_index = (views_index + 1) % len(views_id)
                else:
                    while True:
                        order = (requests.get(
                            f'https://smmfollows.com/api/v2?key=83a67d5b6906fd48f8ee2a7343630835&action=add&service={str(likes_id[likes_index].split(":")[0])}&link={str(link)}&quantity={str(int(likes))}') if 'smm' in str(
                            likes_id[likes_index].split(":")[1]) else requests.get(
                            f'https://paytosmm.com/api/v2?key=20069f39dcdd7aa14cecdb66afd03389&action=add&service={str(likes_id[likes_index].split(":")[0])}&link={str(link)}&quantity={str(int(likes))}')
                                 )
                        response = order.json()
                        if 'error' in response:
                            webpage = str(likes_id[likes_index].split(":")[1])
                            if 'Not enough funds on balance' in response['error']:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order likes for posts due to error : \n{response["error"]}')
                                input(
                                    f'Error Post Likes {webpage} : Please recharge your account and press any key to continue')
                            elif 'less than minimal' in response['error']:
                                get_minimal = int(response['error'].replace('Quantity less than minimal', '').strip())
                                likes = get_minimal
                            else:
                                emailAlert(f'{webpage} Error',
                                           f'Hello Mr Chew\nThe FB bot.\nIt is to inform you that I cant order likes for posts due to error : \n{response["error"]}')
                                print(f'Error Post Likes {webpage}' + response['error'])
                                input(
                                    f'Error Post Likes {webpage} : At id {likes_id[likes_index]}Correct Error and press any key to continue')
                                sheet = load_workbook('fb_list.xlsx')['smm_follows_ids']
                                likes_id = list(_id[0].value for _id in sheet['B2':'B' + self.get_max_row(sheet, 'B')])
                        else:
                            break
                likes_index = (likes_index + 1) % len(likes_id)

    def get_rules(self):
        rules_file = open('rules.json', 'r')
        rules = json.load(rules_file)
        rules_file.close()
        for d in rules:
            _range = d.split('-')
            _from = _range[0]
            _to = _range[1]
            if _to == 'infinity':
                if self.no_of_followers > int(_from):
                    return rules[d]
            elif self.no_of_followers in range(int(_from), int(_to)):
                return rules[d]
        return {}
